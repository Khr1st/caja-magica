"""Generador de Excel con identidad visual Kent / KYNEX Ventures."""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent

# --- Paleta Kent ---
COLOR_HEADER_BG = "1A4731"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ROW_ALT = "F0F7F0"
COLOR_ROW_PLAIN = "FFFFFF"
COLOR_TOTAL_BG = "FFF9C4"
COLOR_TOTAL_FONT = "1A4731"
COLOR_BORDER = "CCCCCC"
COLOR_INPUT_BG = "EBF5FB"
COLOR_INPUT_FONT = "1F4E79"
FONT_NAME = "Trebuchet MS"

COP_FMT = '#,##0" COP"'
USD_FMT = '"$"#,##0.00'

_border_side = Side(style="thin", color=COLOR_BORDER)
THIN_BORDER = Border(left=_border_side, right=_border_side, top=_border_side, bottom=_border_side)


def _header_font():
    return Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_FONT, size=11)


def _header_fill():
    return PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")


def _alt_fill():
    return PatternFill(start_color=COLOR_ROW_ALT, end_color=COLOR_ROW_ALT, fill_type="solid")


def _plain_fill():
    return PatternFill(start_color=COLOR_ROW_PLAIN, end_color=COLOR_ROW_PLAIN, fill_type="solid")


def _total_fill():
    return PatternFill(start_color=COLOR_TOTAL_BG, end_color=COLOR_TOTAL_BG, fill_type="solid")


def _total_font():
    return Font(name=FONT_NAME, bold=True, color=COLOR_TOTAL_FONT, size=11)


def _input_fill():
    return PatternFill(start_color=COLOR_INPUT_BG, end_color=COLOR_INPUT_BG, fill_type="solid")


def _input_font():
    return Font(name=FONT_NAME, color=COLOR_INPUT_FONT, size=11)


def _data_font():
    return Font(name=FONT_NAME, size=10)


def _title_font():
    return Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_BG, size=14)


def _subtitle_font():
    return Font(name=FONT_NAME, color="666666", size=10)


def apply_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_data_row(ws, row, cols, alt=False):
    fill = _alt_fill() if alt else _plain_fill()
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _data_font()
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def apply_total_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _total_font()
        cell.fill = _total_fill()
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_input_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _input_font()
        cell.fill = _input_fill()
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def auto_col_width(ws, min_w=10, max_w=45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 3, min_w), max_w)
        ws.column_dimensions[col_letter].width = width


def _sheet_registro(wb, movimientos, mes):
    ws = wb.active
    ws.title = "Registro"
    cols = 9

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    title_cell = ws.cell(row=1, column=1, value=f"CAJA MÁGICA — Registro {mes}")
    title_cell.font = _title_font()
    title_cell.alignment = Alignment(horizontal="center")

    # Subtítulo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    sub_cell = ws.cell(row=2, column=1, value=f"KYNEX Ventures · Kent Díaz · Exportado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    sub_cell.font = _subtitle_font()
    sub_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Fecha", "Hora", "Descripción", "Tipo", "Categoría", "Moneda", "Monto Original", "Monto COP", "Proyección"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    apply_header(ws, 3, cols)

    # Datos ordenados ASC
    sorted_movs = sorted(movimientos, key=lambda m: m.get("timestamp", ""))
    for i, mov in enumerate(sorted_movs):
        row = 4 + i
        ts = mov.get("timestamp", "")
        fecha = ts[:10] if len(ts) >= 10 else ""
        hora = ts[11:16] if len(ts) >= 16 else ""

        ws.cell(row=row, column=1, value=fecha)
        ws.cell(row=row, column=2, value=hora)
        ws.cell(row=row, column=3, value=mov.get("descripcion", ""))
        ws.cell(row=row, column=4, value=mov.get("tipo", ""))
        ws.cell(row=row, column=5, value=mov.get("categoria", ""))
        ws.cell(row=row, column=6, value=mov.get("moneda", "COP"))

        monto_orig = mov.get("monto_original", 0)
        moneda = mov.get("moneda", "COP")
        orig_cell = ws.cell(row=row, column=7, value=monto_orig)
        if moneda == "USD":
            orig_cell.number_format = USD_FMT
        else:
            orig_cell.number_format = COP_FMT

        cop_cell = ws.cell(row=row, column=8, value=mov.get("monto_cop", 0))
        cop_cell.number_format = COP_FMT

        ws.cell(row=row, column=9, value="Sí" if mov.get("es_proyeccion") else "No")

        apply_data_row(ws, row, cols, alt=(i % 2 == 1))

    # Fila total
    total_row = 4 + len(sorted_movs)
    ws.cell(row=total_row, column=1, value="TOTAL")
    first_data = 4
    last_data = total_row - 1
    if last_data >= first_data:
        ws.cell(row=total_row, column=8).value = f"=SUM(H{first_data}:H{last_data})"
        ws.cell(row=total_row, column=8).number_format = COP_FMT
    apply_total_row(ws, total_row, cols)

    auto_col_width(ws)
    return ws


def _sheet_resumen(wb, movimientos, mes):
    ws = wb.create_sheet("Resumen")

    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=1, value=f"RESUMEN — {mes}").font = _title_font()

    headers = ["Métrica", "Valor"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    apply_header(ws, 3, 2)

    reg_ws = wb["Registro"]
    n_data = len(movimientos)
    first = 4
    last = 3 + n_data if n_data > 0 else 4

    metrics = [
        ("Mes analizado", mes),
        ("Total ingresos COP", f'=SUMIF(Registro!D{first}:D{last},"ingreso",Registro!H{first}:H{last})'),
        ("Total egresos COP", f'=SUMIF(Registro!D{first}:D{last},"egreso",Registro!H{first}:H{last})'),
        ("Total ahorros COP", f'=SUMIF(Registro!D{first}:D{last},"ahorro",Registro!H{first}:H{last})'),
        ("Flujo neto COP", "=B5-B6-B7"),
        ("Movimientos totales", f"=COUNTA(Registro!C{first}:C{last})"),
    ]

    for i, (label, value) in enumerate(metrics):
        row = 4 + i
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if i >= 1 and i <= 4:
            ws.cell(row=row, column=2).number_format = COP_FMT
        apply_data_row(ws, row, 2, alt=(i % 2 == 1))

    auto_col_width(ws)


def _sheet_flujo(wb, mes):
    ws = wb.create_sheet("Flujo 6 Meses")

    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1, value="PROYECCIÓN DE FLUJO — 6 MESES").font = _title_font()

    headers = ["Mes", "Ingresos Proyectados", "Egresos Proyectados", "Neto", "Caja Acumulada"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    apply_header(ws, 3, 5)

    # Parse current month
    try:
        year, month = int(mes[:4]), int(mes[5:7])
    except (ValueError, IndexError):
        year, month = datetime.now().year, datetime.now().month

    for i in range(6):
        row = 4 + i
        m = month + i
        y = year
        while m > 12:
            m -= 12
            y += 1
        mes_label = f"{y}-{m:02d}"

        ws.cell(row=row, column=1, value=mes_label)

        if i == 0:
            # Real data from Resumen
            ws.cell(row=row, column=2, value="=Resumen!B5")
            ws.cell(row=row, column=3, value="=Resumen!B6")
        else:
            # Editable cells
            ws.cell(row=row, column=2, value=0)
            ws.cell(row=row, column=3, value=0)
            apply_input_row(ws, row, 5)

        # Neto = Ingresos - Egresos
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}")

        # Caja Acumulada
        if i == 0:
            ws.cell(row=row, column=5, value=f"=D{row}")
        else:
            ws.cell(row=row, column=5, value=f"=E{row - 1}+D{row}")

        for c in range(2, 6):
            ws.cell(row=row, column=c).number_format = COP_FMT

        if i == 0:
            apply_data_row(ws, row, 5)

    # Note
    note_row = 11
    ws.merge_cells(f"A{note_row}:E{note_row}")
    note = ws.cell(row=note_row, column=1, value="Las celdas en azul claro son editables — ingresa tus estimaciones")
    note.font = Font(name=FONT_NAME, italic=True, color="888888", size=9)

    auto_col_width(ws)


def _sheet_supuestos(wb):
    ws = wb.create_sheet("Supuestos")

    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=1, value="SUPUESTOS Y PARÁMETROS").font = _title_font()

    headers = ["Parámetro", "Valor"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    apply_header(ws, 3, 2)

    params = [
        ("Tasa COP/USD", 4200),
        ("Precio hora consultoría (USD)", 25),
        ("Horas facturables/semana (máx real: 10h)", 10),
        ("Caja mínima COP (umbral de alerta)", 800000),
        ("% reinversión primeros ingresos (regla 50/50)", 50),
        ("Objetivo ingreso mensual USD (meta Oklahoma)", 2000),
    ]

    for i, (label, value) in enumerate(params):
        row = 4 + i
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        apply_input_row(ws, row, 2)

    auto_col_width(ws)


def generar_excel(movimientos: list, mes: str) -> str:
    """Genera Excel con 4 hojas y devuelve la ruta del archivo."""
    wb = Workbook()

    _sheet_registro(wb, movimientos, mes)
    _sheet_resumen(wb, movimientos, mes)
    _sheet_flujo(wb, mes)
    _sheet_supuestos(wb)

    out_dir = BASE_DIR / "data"
    out_dir.mkdir(exist_ok=True)
    path = out_dir / f"CajaMagica_{mes}.xlsx"
    wb.save(str(path))
    return str(path)
